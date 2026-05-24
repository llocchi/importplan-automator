import base64, io
from datetime import datetime
from typing import Dict, Tuple
import openpyxl
from logger_setup import get_logger

log = get_logger('tuut.xlsx')


def _norm(val):
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def update_xlsx(xlsx_base64: str, triplas: Dict[Tuple[str,str],str],
                encontrados: Dict[Tuple[str,str],int], original_name: str) -> dict:
    try:
        xlsx_bytes = base64.b64decode(xlsx_base64)
    except Exception:
        raise ValueError('xlsx_base64 invalido')

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    log.info('Sheets no XLSX: %s | Ativa: %s', wb.sheetnames, wb.active.title)
    ws = wb.active
    log.info('XLSX: sheet=%s max_row=%d max_col=%d', ws.title, ws.max_row, ws.max_column)

    header = [_norm(ws.cell(1, c).value) for c in range(1, min(7, ws.max_column + 1))]
    log.info('XLSX cabecalho primeiras colunas: %s', header)

    # --- 1. Monta indice XLSX: (SIS, REF) -> lista de (row_num, seq_cell) ---
    xlsx_index: Dict[Tuple[str,str], list] = {}
    total_xlsx_linhas = 0
    for row in ws.iter_rows(min_row=2):
        if len(row) < 4:
            continue
        s = _norm(row[2].value)
        r = _norm(row[3].value)
        if not s or not r or s in ('None', 'SISTEMA') or r in ('None', 'REF.'):
            continue
        total_xlsx_linhas += 1
        key = (s, r)
        if key not in xlsx_index:
            xlsx_index[key] = []
        xlsx_index[key].append((row[0].row, row[0]))

    log.info('XLSX index: %d pares unicos | %d linhas com SIS+REF validos | %d no PDF',
             len(xlsx_index), total_xlsx_linhas, len(triplas))

    # --- 2. Indice sis-only do PDF para fallback ---
    triplas_by_sis: Dict[str, tuple] = {}
    sis_to_ref_pdf: Dict[str, str] = {}
    for (s, r), sq in triplas.items():
        if s not in triplas_by_sis:
            triplas_by_sis[s] = (sq, encontrados.get((s, r), '?'))
            sis_to_ref_pdf[s] = r

    # --- 3. Itera linhas do XLSX e preenche com seq do PDF ---
    updated = 0
    not_found_list = []
    exact_count = 0
    fallback_count = 0
    not_found_count = 0

    for (sis, ref), xlsx_rows in xlsx_index.items():
        seq = None
        pag = '?'
        caminho = None
        if (sis, ref) in triplas:
            seq = triplas[(sis, ref)]
            pag = encontrados.get((sis, ref), '?')
            caminho = 'exato'
            exact_count += 1
        elif sis in triplas_by_sis:
            seq, pag = triplas_by_sis[sis]
            caminho = 'fallback'
            fallback_count += 1
            ref_pdf = sis_to_ref_pdf.get(sis, '?')
            log.warning('FALLBACK sis=%s ref_xlsx=%s ref_pdf=%s seq=%s pag=%s',
                        sis, ref, ref_pdf, seq, pag)

        if seq is not None:
            for (row_num, seq_cell) in xlsx_rows:
                try:
                    seq_cell.value = int(seq)
                except ValueError:
                    seq_cell.value = seq
                updated += 1
                log.info('UPDATE seq=%s sis=%s ref=%s pag=%s via=%s row=%d',
                         seq, sis, ref, pag, caminho, row_num)
        else:
            not_found_count += 1
            not_found_list.append({'sistema': sis, 'ref': ref, 'motivo': 'nao no PDF'})
            log.error('NAO_ENCONTRADO sis=%s ref=%s', sis, ref)

    log.info('XLSX FIM: pares_unicos=%d linhas_xlsx=%d pdf_produtos=%d exato=%d fallback=%d nao_encontrado=%d rows_atualizadas=%d',
             len(xlsx_index), total_xlsx_linhas, len(triplas),
             exact_count, fallback_count, not_found_count, updated)

    if updated == 0:
        log.error('ZERO linhas atualizadas! Veja diagnostico acima.')

    base_name = original_name
    for ext in ('.xlsx', '.xls', '.XLSX', '.XLS'):
        base_name = base_name.replace(ext, '')
    ts = datetime.now().strftime('%d-%m-%Y_%H%M%S')
    output_filename = f'{base_name}_atualizado_{ts}.xlsx'
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    b64 = base64.b64encode(out.read()).decode()
    return {
        'xlsx_base64': b64,
        'report': {
            'total_rows': total_xlsx_linhas,
            'updated': updated,
            'not_found': not_found_list,
            'output_filename': output_filename,
            'processed_at': datetime.now().isoformat(),
        }
    }
